"""
Stage3 Preprocessing Module
---------------------------

This module performs the following operations:

1. Remove "未着手" rows from progress data.
2. Normalize repair task names (補修①〜⑩ → 補修1〜補修10).
3. Add a combined key column to history data.
4. Clean newline characters from specific columns.
5. Prevent scientific notation in Excel output.

This script is intended to be used as part of the Stage3 pipeline
in the Data Preprocessing Pipeline project.

Author: YOUR_NAME
Repository: https://github.com/yourname/yourrepo
"""

from pathlib import Path
import time
import pandas as pd
import openpyxl


# ==============================================================================
# 補修①〜⑩ 名称変換
# ==============================================================================

def process_repair_names(df_progress: pd.DataFrame,
                         df_kiki: pd.DataFrame) -> pd.DataFrame:
    """
    Replace repair task names (補修①〜⑩) with normalized names (補修1〜補修10).

    Parameters
    ----------
    df_progress : pd.DataFrame
        Progress sheet data.
    df_kiki : pd.DataFrame
        Equipment/Construction sheet data.

    Returns
    -------
    pd.DataFrame
        Updated progress DataFrame.
    """
    repair_map = {
        "補修①": "補修1", "補修②": "補修2", "補修③": "補修3",
        "補修④": "補修4", "補修⑤": "補修5", "補修⑥": "補修6",
        "補修⑦": "補修7", "補修⑧": "補修8", "補修⑨": "補修9",
        "補修⑩": "補修10",
    }

    mask = df_progress["作業名"].isin(repair_map.keys())
    df_target = df_progress[mask].copy()

    if df_target.empty:
        print("INFO: 補修①〜⑩ の作業名は見つかりませんでした。")
        return df_progress

    df_target["補修列名"] = df_target["作業名"].map(repair_map)

    df_merged = df_target.merge(
        df_kiki,
        on="機器・工事ID",
        how="left",
        suffixes=("", "_kiki")
    )

    df_merged["新作業名"] = df_merged.apply(
        lambda row: row.get(row["補修列名"], row["作業名"]),
        axis=1
    )

    df_progress.loc[mask, "作業名"] = df_merged["新作業名"].values
    return df_progress


# ==============================================================================
# メイン処理
# ==============================================================================

def load_latest_excel(input_dir: Path) -> Path:
    """Return the most recently updated Excel file in the directory."""
    files = sorted(input_dir.glob("*.xlsx"),
                   key=lambda f: f.stat().st_mtime,
                   reverse=True)

    if not files:
        raise FileNotFoundError("ERROR: .xlsx ファイルが見つかりません。")

    return files[0]


def clean_newlines(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Remove newline characters from specified columns."""
    for col in columns:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(r"[\r\n]+", "", regex=True)
        else:
            print(f"WARNING: 列『{col}』が見つかりませんでした。")
    return df


def add_history_key(df_history: pd.DataFrame) -> pd.DataFrame:
    """Add 作業＆機器工事ID column to history sheet."""
    new_col = "作業＆機器工事ID"

    if new_col not in df_history.columns:
        insert_pos = df_history.columns.get_loc("実績") + 1 \
            if "実績" in df_history.columns else len(df_history.columns)

        df_history.insert(
            insert_pos,
            new_col,
            df_history["作業ID"].astype(str) + "_" +
            df_history["機器・工事ID"].astype(str)
        )
    else:
        print(f"INFO: 列『{new_col}』は既に存在します。")

    return df_history


def fix_scientific_notation(ws, col_name: str):
    """Force Excel column to treat values as text."""
    excel_cols = {cell.value: cell.column for cell in ws[1]}

    if col_name not in excel_cols:
        print(f"WARNING: Excel に列『{col_name}』が見つかりません。")
        return

    col_idx = excel_cols[col_name]

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.data_type = "s"
        cell.number_format = "@"
        cell.value = "'" + str(cell.value)


def main():
    """Main execution flow for Stage3 preprocessing."""
    base_dir = Path.cwd()
    input_dir = base_dir.parent / "ファイル設置場所"

    excel_path = load_latest_excel(input_dir)
    print(f"INFO: 対象ファイル → {excel_path.name}")

    df_progress = pd.read_excel(excel_path, sheet_name="進捗情報")
    df_kiki = pd.read_excel(excel_path, sheet_name="機器工事情報")
    df_history = pd.read_excel(excel_path, sheet_name="進捗履歴情報")

    time.sleep(1)

    # 機器工事情報：機器・工事No の先頭 ' を除去
    if "機器・工事No" in df_kiki.columns:
        df_kiki["機器・工事No"] = df_kiki["機器・工事No"].astype(str).str.strip("'")

    # 未着手削除
    df_progress = df_progress[
        ~((df_progress["進捗"] == "未着手") | (df_progress["実績"] == "未着手"))
    ]

    # 補修①〜⑩ 名称置換
    df_progress = process_repair_names(df_progress, df_kiki)

    # 履歴情報：キー追加
    df_history = add_history_key(df_history)

    # 改行除去
    df_progress = clean_newlines(
        df_progress,
        ["機器・工事名", "機器・工事名称", "会社名", "プラント名"]
    )

    # Excel 書き込み
    with pd.ExcelWriter(excel_path, engine="openpyxl",
                        mode="a", if_sheet_exists="replace") as writer:
        df_progress.to_excel(writer, sheet_name="進捗情報", index=False)
        df_history.to_excel(writer, sheet_name="進捗履歴情報", index=False)
        df_kiki.to_excel(writer, sheet_name="機器工事情報", index=False)

    # 指数表記対策
    wb = openpyxl.load_workbook(excel_path)
    fix_scientific_notation(wb["進捗情報"], "機器・工事No")
    fix_scientific_notation(wb["機器工事情報"], "機器・工事No")
    wb.save(excel_path)

    print("INFO: Stage3 preprocessing completed.")


if __name__ == "__main__":
    main()
